import io

import pytest
from openpyxl import Workbook

from app import planilha

# Cabecalho exatamente como o scraper do Google Maps escreve.
CABECALHO = [
    "Nome",
    "Telefone",
    "Endereco",
    "Categoria",
    "Nota",
    "Avaliacoes",
    "Anuncio",
    "Link Maps",
    "Busca",
    "Capturado em",
]

# Lojas reais capturadas em Canoas/RS. Duas delas sao fixo - proporcao parecida
# com a que o Maps devolve de verdade.
BICHO_MANIA = [
    "Bicho Mania", "(51) 99898-4086", "R. Joaquim Caetano, 211", "Banho e tosa",
    4.8, 312, "Nao", "https://maps.google.com/?cid=1", "pet shop canoas",
    "2026-07-20 09:10",
]
CLINICA_DAOIA = [
    "Clinica Veterinaria Dra. Daoia Taine", "(51) 3466-0454", "R. Rui Barbosa, 351",
    "Pet Shop", 4.9, 87, "Nao", "https://maps.google.com/?cid=2", "pet shop canoas",
    "2026-07-20 09:10",
]
PETZ = [
    "Petz Canoas", "(51) 3052-0478", "Av. Getulio Vargas, 6401", "Pet Shop",
    4.5, 1204, "Sim", "https://maps.google.com/?cid=3", "pet shop canoas",
    "2026-07-20 09:10",
]
AGROPET = [
    "Agropet Tipo Bicho", "(51) 99858-1025", "R. Joaquim Nabuco, 171", "Pet Shop",
    4.7, 45, "Nao", "https://maps.google.com/?cid=4", "pet shop canoas",
    "2026-07-20 09:10",
]
CAOTELLI = [
    "CaoTelli", "(51) 99765-5755", "Av. Santos Ferreira, 997", "Pet Shop",
    5.0, 23, "Nao", "https://maps.google.com/?cid=5", "pet shop canoas",
    "2026-07-20 09:10",
]

TODAS = [BICHO_MANIA, CLINICA_DAOIA, PETZ, AGROPET, CAOTELLI]


def montar_xlsx(tmp_path, linhas, cabecalho=None, nome="lojas.xlsx"):
    """Gera a planilha do scraper em disco, dentro do tmp_path do pytest."""
    livro = Workbook()
    aba = livro.active
    aba.append(list(cabecalho if cabecalho is not None else CABECALHO))
    for linha in linhas:
        aba.append(list(linha))
    caminho = tmp_path / nome
    livro.save(caminho)
    return str(caminho)


def montar_bytes(linhas, cabecalho=None) -> io.BytesIO:
    """Mesma planilha, mas em memoria - e assim que o upload chega."""
    livro = Workbook()
    aba = livro.active
    aba.append(list(cabecalho if cabecalho is not None else CABECALHO))
    for linha in linhas:
        aba.append(list(linha))
    buffer = io.BytesIO()
    livro.save(buffer)
    buffer.seek(0)
    return buffer


def reordenar(cabecalho, linha, ordem):
    """Reescreve uma linha na ordem de colunas pedida."""
    posicao = {rotulo: i for i, rotulo in enumerate(cabecalho)}
    return [linha[posicao[rotulo]] for rotulo in ordem]


class TestCabecalho:
    def test_le_o_cabecalho_do_scraper(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, TODAS))
        assert r.total_linhas == 5
        assert r.prontos == 3

    def test_ordem_das_colunas_nao_importa(self, tmp_path):
        ordem = ["Telefone", "Busca", "Nome", "Link Maps", "Categoria", "Endereco"]
        linhas = [reordenar(CABECALHO, l, ordem) for l in TODAS]
        r = planilha.ler(montar_xlsx(tmp_path, linhas, cabecalho=ordem))
        assert r.prontos == 3
        assert r.leads[0].nome == "Bicho Mania"
        assert r.leads[0].endereco == "R. Joaquim Caetano, 211"

    def test_cabecalho_com_acento_e_aceito(self, tmp_path):
        # O usuario abriu no Excel e "corrigiu" a grafia.
        cabecalho = list(CABECALHO)
        cabecalho[2] = "Endereço"
        cabecalho[5] = "Avaliações"
        r = planilha.ler(montar_xlsx(tmp_path, TODAS, cabecalho=cabecalho))
        assert r.leads[0].endereco == "R. Joaquim Caetano, 211"

    def test_caixa_e_espaco_sobrando_nao_atrapalham(self, tmp_path):
        cabecalho = ["  nome ", "TELEFONE", "endereco", "  Categoria",
                     "Nota", "Avaliacoes", "Anuncio", "link maps", "BUSCA",
                     "Capturado Em"]
        r = planilha.ler(montar_xlsx(tmp_path, TODAS, cabecalho=cabecalho))
        assert r.prontos == 3
        assert r.leads[0].categoria == "Banho e tosa"

    def test_coluna_desconhecida_e_ignorada(self, tmp_path):
        # Anotacao propria do usuario nao pode quebrar a leitura.
        cabecalho = CABECALHO + ["Meu comentario"]
        linhas = [l + ["ligar depois"] for l in TODAS]
        r = planilha.ler(montar_xlsx(tmp_path, linhas, cabecalho=cabecalho))
        assert r.prontos == 3

    def test_sem_coluna_telefone_o_erro_diz_o_que_faltou(self, tmp_path):
        cabecalho = [c for c in CABECALHO if c != "Telefone"]
        linhas = [reordenar(CABECALHO, l, cabecalho) for l in TODAS]
        caminho = montar_xlsx(tmp_path, linhas, cabecalho=cabecalho)

        with pytest.raises(planilha.PlanilhaInvalida) as erro:
            planilha.ler(caminho)

        texto = str(erro.value)
        assert "telefone" in texto.lower()
        assert "Nome" in texto, "o erro precisa listar os cabecalhos encontrados"

    def test_sem_coluna_nome_tambem_falha(self, tmp_path):
        cabecalho = [c for c in CABECALHO if c != "Nome"]
        linhas = [reordenar(CABECALHO, l, cabecalho) for l in TODAS]
        with pytest.raises(planilha.PlanilhaInvalida, match="(?i)nome"):
            planilha.ler(montar_xlsx(tmp_path, linhas, cabecalho=cabecalho))

    def test_planilha_totalmente_vazia_falha_com_mensagem_clara(self, tmp_path):
        livro = Workbook()
        caminho = tmp_path / "vazia.xlsx"
        livro.save(caminho)
        with pytest.raises(planilha.PlanilhaInvalida, match="(?i)vazia"):
            planilha.ler(str(caminho))

    def test_arquivo_que_nao_e_xlsx_falha_sem_estourar_stacktrace(self, tmp_path):
        caminho = tmp_path / "lista.txt"
        caminho.write_text("Bicho Mania;(51) 99898-4086", encoding="utf-8")
        with pytest.raises(planilha.PlanilhaInvalida):
            planilha.ler(str(caminho))


class TestNormalizacaoDeCabecalho:
    @pytest.mark.parametrize(
        "bruto",
        ["Endereço", "  ENDERECO  ", "endereco", "Endereco", "ENDEREÇO"],
    )
    def test_variacoes_do_mesmo_rotulo_batem(self, bruto):
        assert planilha.normalizar_cabecalho(bruto) == "endereco"

    def test_underline_vira_espaco(self):
        assert planilha.normalizar_cabecalho("Link_Maps") == "link maps"


class TestClassificacaoDasLinhas:
    def test_fixo_e_descartado(self, tmp_path):
        # Clinica Dra. Daoia Taine e Petz sao fixo: quase nunca tem WhatsApp.
        r = planilha.ler(montar_xlsx(tmp_path, TODAS))
        assert r.fixos == 2
        assert all("Petz" not in lead.nome for lead in r.leads)
        assert all("Daoia" not in lead.nome for lead in r.leads)

    def test_linha_sem_telefone_nao_conta_como_erro_de_leitura(self, tmp_path):
        sem_fone = list(BICHO_MANIA)
        sem_fone[0] = "Pet sem telefone"
        sem_fone[1] = None
        r = planilha.ler(montar_xlsx(tmp_path, [sem_fone, AGROPET]))
        assert r.total_linhas == 2
        assert r.sem_telefone == 1
        assert r.prontos == 1

    def test_telefone_ilegivel_entra_como_invalido(self, tmp_path):
        lixo = list(CAOTELLI)
        lixo[0] = "Pet do zap errado"
        lixo[1] = "chama no direct"
        r = planilha.ler(montar_xlsx(tmp_path, [lixo, AGROPET]))
        assert r.invalidos == 1
        assert r.prontos == 1

    def test_telefone_guardado_como_numero_pelo_excel(self, tmp_path):
        # Celula formatada como numero volta como float e nao pode virar
        # "51998581025.0".
        numerico = list(AGROPET)
        numerico[1] = 51998581025
        r = planilha.ler(montar_xlsx(tmp_path, [numerico]))
        assert r.prontos == 1
        assert r.leads[0].telefone == "5551998581025"

    def test_celular_vira_e164_e_exibicao_legivel(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA]))
        lead = r.leads[0]
        assert lead.telefone == "5551998984086"
        assert lead.telefone_exibicao == "(51) 99898-4086"


class TestDeduplicacaoInterna:
    def test_mesma_loja_duas_vezes_entra_uma_vez(self, tmp_path):
        # O scraper repete quando a loja aparece em duas buscas diferentes.
        outra_busca = list(BICHO_MANIA)
        outra_busca[8] = "banho e tosa canoas"
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA, outra_busca, AGROPET]))
        assert r.duplicados == 1
        assert r.prontos == 2

    def test_duplicata_e_pelo_telefone_normalizado(self, tmp_path):
        # Mesmo numero escrito de outro jeito continua sendo o mesmo numero.
        disfarcado = list(BICHO_MANIA)
        disfarcado[0] = "Bicho Mania - Matriz"
        disfarcado[1] = "+55 51 99898 4086"
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA, disfarcado]))
        assert r.duplicados == 1
        assert r.prontos == 1

    def test_mantem_a_primeira_ocorrencia(self, tmp_path):
        segunda = list(BICHO_MANIA)
        segunda[0] = "Bicho Mania (filial)"
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA, segunda]))
        assert r.leads[0].nome == "Bicho Mania"

    def test_lojas_diferentes_nao_sao_duplicata(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA, AGROPET, CAOTELLI]))
        assert r.duplicados == 0
        assert r.prontos == 3


class TestLinhasVazias:
    def test_linhas_vazias_no_fim_sao_ignoradas(self, tmp_path):
        vazias = [[None] * len(CABECALHO) for _ in range(3)]
        r = planilha.ler(montar_xlsx(tmp_path, TODAS + vazias))
        assert r.total_linhas == 5, "linha vazia do xlsx nao e loja"
        assert r.prontos == 3

    def test_linhas_com_celulas_em_branco_tambem_sao_ignoradas(self, tmp_path):
        brancas = [[""] * len(CABECALHO), ["   "] * len(CABECALHO)]
        r = planilha.ler(montar_xlsx(tmp_path, TODAS + brancas))
        assert r.total_linhas == 5

    def test_linha_vazia_no_meio_nao_interrompe_a_leitura(self, tmp_path):
        linhas = [BICHO_MANIA, [None] * len(CABECALHO), AGROPET]
        r = planilha.ler(montar_xlsx(tmp_path, linhas))
        assert r.total_linhas == 2
        assert r.prontos == 2


class TestRelatorio:
    def test_a_conta_fecha(self, tmp_path):
        # Sem esta invariante o usuario nao consegue explicar as lojas que
        # sumiram entre o arquivo e a fila.
        sem_fone = list(CAOTELLI)
        sem_fone[1] = None
        lixo = list(CAOTELLI)
        lixo[1] = "sem numero"
        linhas = TODAS + [BICHO_MANIA, sem_fone, lixo, [None] * len(CABECALHO)]

        r = planilha.ler(montar_xlsx(tmp_path, linhas))

        assert r.total_linhas == 8
        assert (
            r.sem_telefone + r.invalidos + r.duplicados + r.fixos + r.prontos
            == r.total_linhas
        )

    def test_contagens_da_tela(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, TODAS))
        assert r.total_linhas == 5
        assert r.com_telefone == 5
        assert r.prontos == 3
        assert r.fixos == 2
        assert r.descartados == 2

    def test_resumo_em_uma_linha(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, TODAS))
        assert r.resumo() == (
            "5 lojas lidas, 5 com telefone, 3 prontas para disparo, 2 sao fixo"
        )

    def test_detalhes_so_mostram_o_que_aconteceu(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA, AGROPET]))
        assert r.detalhes() == []

    def test_detalhes_explicam_cada_descarte(self, tmp_path):
        lixo = list(CAOTELLI)
        lixo[1] = "manda dm"
        r = planilha.ler(montar_xlsx(tmp_path, TODAS + [lixo]))
        texto = " ".join(r.detalhes())
        assert "fixo" in texto
        assert "nao da para ler" in texto

    def test_planilha_so_com_cabecalho_nao_e_erro(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, []))
        assert r.total_linhas == 0
        assert r.leads == ()


class TestOrigemDoArquivo:
    def test_aceita_caminho_no_disco(self, tmp_path):
        assert planilha.ler(montar_xlsx(tmp_path, TODAS)).prontos == 3

    def test_aceita_upload_em_memoria(self):
        assert planilha.ler(montar_bytes(TODAS)).prontos == 3

    def test_upload_e_disco_dao_o_mesmo_resultado(self, tmp_path):
        do_disco = planilha.ler(montar_xlsx(tmp_path, TODAS))
        da_memoria = planilha.ler(montar_bytes(TODAS))
        assert do_disco == da_memoria


class TestCamposDoLead:
    def test_lead_carrega_o_que_a_mensagem_precisa(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA]))
        lead = r.leads[0]
        assert lead.nome == "Bicho Mania"
        assert lead.endereco == "R. Joaquim Caetano, 211"
        assert lead.categoria == "Banho e tosa"
        assert lead.busca == "pet shop canoas"
        assert lead.link_maps == "https://maps.google.com/?cid=1"

    def test_coluna_opcional_ausente_vira_string_vazia(self, tmp_path):
        # Nunca None: a mensagem nao pode sair com "None" no meio.
        cabecalho = ["Nome", "Telefone"]
        linhas = [[BICHO_MANIA[0], BICHO_MANIA[1]]]
        r = planilha.ler(montar_xlsx(tmp_path, linhas, cabecalho=cabecalho))
        lead = r.leads[0]
        assert lead.categoria == ""
        assert lead.endereco == ""
        assert lead.busca == ""

    def test_lead_e_imutavel(self, tmp_path):
        r = planilha.ler(montar_xlsx(tmp_path, [BICHO_MANIA]))
        with pytest.raises(Exception):
            r.leads[0].nome = "outro"
