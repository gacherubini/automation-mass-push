"""Leitura da planilha de lojas que o scraper do Google Maps gera.

O arquivo chega de duas formas: caminho no disco (uso local) ou upload no
dashboard (BytesIO). As duas entram pela mesma porta - `ler()`.

Duas coisas justificam este modulo existir separado:

1. A planilha passa pela mao do usuario antes de voltar para ca. Ele reordena
   coluna, renomeia "Endereco" para "Endereco" com acento, apaga o que nao usa.
   Por isso o mapeamento e feito PELO CABECALHO, comparando uma forma
   normalizada (sem acento, sem caixa, sem espaco sobrando), nunca por posicao.

2. O que sai daqui nao e "as linhas do arquivo", e um relatorio: quantas lojas
   entraram, quantas sobraram e por que as outras cairam. Essa contagem e o que
   a tela mostra antes de o usuario apertar "disparar" - e a ultima chance de
   ele perceber que a planilha veio errada.

Funcao pura no que importa: a unica I/O e abrir o arquivo que foi entregue.
"""

from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app import telefone as tel

# Nome canonico de cada coluna -> formas aceitas no cabecalho, ja normalizadas.
# O usuario pode ter mexido no arquivo; o scraper nao e o unico a escrever nele.
ALIASES_DE_COLUNA: dict[str, tuple[str, ...]] = {
    "nome": ("nome", "loja", "empresa", "nome da loja"),
    "telefone": ("telefone", "fone", "celular", "whatsapp"),
    "endereco": ("endereco", "endereco completo", "local"),
    "categoria": ("categoria", "segmento", "ramo"),
    "nota": ("nota", "avaliacao media", "rating"),
    "avaliacoes": ("avaliacoes", "numero de avaliacoes", "reviews"),
    "anuncio": ("anuncio", "patrocinado", "ads"),
    "link_maps": ("link maps", "link do maps", "link", "maps", "url"),
    "busca": ("busca", "termo de busca", "pesquisa"),
    "capturado_em": ("capturado em", "data de captura", "capturado"),
}

# Sem estas duas nao ha campanha: uma identifica a loja, a outra e o destino.
COLUNAS_ESSENCIAIS: tuple[str, ...] = ("nome", "telefone")


class PlanilhaInvalida(ValueError):
    """A planilha nao tem o formato minimo para virar campanha."""


@dataclass(frozen=True)
class Lead:
    """Uma loja pronta para receber a primeira mensagem.

    `telefone` e o E.164 sem "+" (vira JID no envio); `telefone_exibicao` e o
    mesmo numero no formato que o usuario reconhece na tela.
    """

    nome: str
    telefone: str
    telefone_exibicao: str
    endereco: str = ""
    categoria: str = ""
    busca: str = ""
    link_maps: str = ""


@dataclass(frozen=True)
class Relatorio:
    """O que a tela mostra depois de ler o arquivo.

    A contagem fecha: total_linhas = sem_telefone + invalidos + duplicados +
    fixos + prontos. Cada linha lida cai em exatamente um balde, senao o
    usuario nao consegue explicar para onde foram as lojas que sumiram.
    """

    total_linhas: int = 0
    sem_telefone: int = 0
    invalidos: int = 0
    duplicados: int = 0
    fixos: int = 0
    leads: tuple[Lead, ...] = field(default_factory=tuple)

    @property
    def prontos(self) -> int:
        return len(self.leads)

    @property
    def com_telefone(self) -> int:
        return self.total_linhas - self.sem_telefone

    @property
    def descartados(self) -> int:
        return self.total_linhas - self.prontos

    def resumo(self) -> str:
        """Uma linha para o topo da tela."""
        return (
            f"{self.total_linhas} lojas lidas, {self.com_telefone} com telefone, "
            f"{self.prontos} prontas para disparo, {self.fixos} sao fixo"
        )

    def detalhes(self) -> list[str]:
        """Motivo de cada descarte, so os que aconteceram."""
        linhas: list[str] = []
        if self.sem_telefone:
            linhas.append(f"{self.sem_telefone} sem telefone na planilha")
        if self.invalidos:
            linhas.append(f"{self.invalidos} com telefone que nao da para ler")
        if self.duplicados:
            linhas.append(f"{self.duplicados} repetidas dentro da propria planilha")
        if self.fixos:
            linhas.append(f"{self.fixos} com telefone fixo (raramente tem WhatsApp)")
        return linhas


def normalizar_cabecalho(texto: Any) -> str:
    """Forma comparavel de um cabecalho: sem acento, sem caixa, sem sobra.

    "  Endereço " e "endereco" precisam bater, senao a planilha salva no Excel
    do usuario deixa de ser lida por causa de um espaco.
    """
    bruto = _texto(texto).replace("_", " ").replace("-", " ")
    sem_acento = "".join(
        c for c in unicodedata.normalize("NFKD", bruto) if not unicodedata.combining(c)
    )
    return " ".join(sem_acento.lower().split())


def mapear_colunas(cabecalho: tuple[Any, ...]) -> dict[str, int]:
    """Descobre em que posicao cada coluna conhecida caiu.

    Colunas desconhecidas sao ignoradas de proposito: o usuario tem direito de
    manter anotacoes proprias na planilha.
    """
    por_alias: dict[str, str] = {
        alias: canonico
        for canonico, aliases in ALIASES_DE_COLUNA.items()
        for alias in aliases
    }

    indices: dict[str, int] = {}
    encontrados: list[str] = []
    for posicao, celula in enumerate(cabecalho):
        rotulo = _texto(celula)
        if not rotulo:
            continue
        encontrados.append(rotulo)
        canonico = por_alias.get(normalizar_cabecalho(rotulo))
        # Cabecalho repetido: a primeira ocorrencia vence, e a da esquerda.
        if canonico and canonico not in indices:
            indices[canonico] = posicao

    faltando = [c for c in COLUNAS_ESSENCIAIS if c not in indices]
    if faltando:
        vistos = ", ".join(encontrados) if encontrados else "nenhum"
        raise PlanilhaInvalida(
            f"Coluna obrigatoria ausente na planilha: {', '.join(faltando)}. "
            f"Cabecalhos encontrados: {vistos}. "
            f"Esperado pelo menos: {', '.join(COLUNAS_ESSENCIAIS)}."
        )

    return indices


def ler(origem: str | BinaryIO) -> Relatorio:
    """Le o .xlsx e devolve o relatorio da campanha.

    `origem` pode ser um caminho ou um objeto de arquivo (o upload do
    dashboard chega como BytesIO).
    """
    try:
        planilha = load_workbook(origem, read_only=True, data_only=True)
    except PlanilhaInvalida:
        raise
    except Exception as erro:  # openpyxl levanta varios tipos para arquivo ruim
        raise PlanilhaInvalida(
            "Nao foi possivel abrir o arquivo. Envie a planilha .xlsx gerada "
            f"pelo scraper do Google Maps. Detalhe tecnico: {erro}"
        ) from erro

    try:
        aba = planilha.active
        linhas = aba.iter_rows(values_only=True) if aba is not None else iter(())

        cabecalho = next(linhas, None)
        if cabecalho is None or _linha_vazia(cabecalho):
            raise PlanilhaInvalida(
                "A planilha esta vazia: nao ha nem a linha de cabecalho com "
                f"{', '.join(COLUNAS_ESSENCIAIS)}."
            )

        indices = mapear_colunas(cabecalho)
        return _consolidar(linhas, indices)
    finally:
        planilha.close()


def _consolidar(linhas, indices: dict[str, int]) -> Relatorio:
    """Classifica cada linha em um unico balde e monta o relatorio."""
    total = sem_telefone = invalidos = duplicados = fixos = 0
    vistos: set[str] = set()
    leads: list[Lead] = []

    for linha in linhas:
        # O xlsx quase sempre traz linhas vazias no fim (o Excel estica a area
        # usada quando alguem clica numa celula la embaixo). Nao e erro.
        if _linha_vazia(linha):
            continue

        total += 1
        bruto = _campo(linha, indices, "telefone")
        if not bruto:
            sem_telefone += 1
            continue

        numero = tel.normalizar(bruto)
        if not numero.valido:
            invalidos += 1
            continue

        # A deduplicacao vem antes da checagem de fixo de proposito: a mesma
        # loja repetida conta como repetida uma vez so, seja qual for o tipo do
        # telefone. Assim os baldes nao se sobrepoem e a soma fecha.
        if numero.numero in vistos:
            duplicados += 1
            continue
        vistos.add(numero.numero)

        if not numero.provavel_whatsapp:
            fixos += 1
            continue

        leads.append(
            Lead(
                nome=_campo(linha, indices, "nome"),
                telefone=numero.numero,
                telefone_exibicao=tel.formatar_br(numero),
                endereco=_campo(linha, indices, "endereco"),
                categoria=_campo(linha, indices, "categoria"),
                busca=_campo(linha, indices, "busca"),
                link_maps=_campo(linha, indices, "link_maps"),
            )
        )

    return Relatorio(
        total_linhas=total,
        sem_telefone=sem_telefone,
        invalidos=invalidos,
        duplicados=duplicados,
        fixos=fixos,
        leads=tuple(leads),
    )


def _campo(linha: tuple[Any, ...], indices: dict[str, int], nome: str) -> str:
    posicao = indices.get(nome)
    if posicao is None or posicao >= len(linha):
        return ""
    return _texto(linha[posicao])


def _linha_vazia(linha: tuple[Any, ...] | None) -> bool:
    return linha is None or all(not _texto(celula) for celula in linha)


def _texto(valor: Any) -> str:
    """Celula do Excel para string limpa.

    Telefone digitado sem mascara chega como numero, e o Excel guarda numero
    como float: 5199898406.0 viraria um telefone com ".0" no fim.
    """
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return "sim" if valor else "nao"
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()
